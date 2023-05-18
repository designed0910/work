#只针对sheet1/支持xls、xlsx/最后输出xlsx格式总表
#可以修改文件价名称/汇总xls、xlsx表格
import xlrd
import os
import openpyxl
import tkinter as tk
import tkinter.filedialog
import re

root_files_Global = []
row = []
column = []
value = []
f0 = 0
 
def askfile():#选择文件夹
    global filename
    filename = tk.filedialog.askdirectory()
    if filename !='':
        label2.config(text=filename)
        print(filename)
    else:
        label2.config(text='您没有选择任何文件夹')

def submit_keyword():
    global keyword_entry, keyword
    keyword = keyword_entry.get().strip()
    # 在这里执行需要用到关键字的代码
    
def GUI_excel():
    global label2, keyword_entry
    window = tk.Tk()
    window.geometry('600x300')
    window.title('excel合成')
    label1 = tk.Label(window,text='请选择需要处理的文件夹，选择好就已经处理完毕，如需重新处理，删除合成好的表并重新启动程序')
    label1.place(x=15,y=10)
    label2 = tk.Label(window,text='')
    label2.place(x=15,y=30)
    button1 = tk.Button(window,text='选择文件夹',bg='blue',fg='white',command=askfile)
    button1.place(x=15,y=60)
    keyword_label = tk.Label(window, text='请输入关键字：')
    keyword_label.place(x=15, y=90)
    keyword_entry = tk.Entry(window)
    keyword_entry.place(x=120, y=90)
    submit_button = tk.Button(window, text='提交', bg='blue', fg='white', command=submit_keyword)
    submit_button.place(x=250, y=85)

    window.mainloop()
 
GUI_excel()



def find_excel(root_folder, keyword):
    
    for root, dirs, files in os.walk(root_folder):
        for file_name in files:
            if re.search(keyword, file_name):
                file_path = os.path.join(root, file_name)
                root_files_Global.append(file_path)
    


find_excel(filename,keyword)
print(root_files_Global)  
    

def read_excel_xls():#定义读取xls文件格式函数
    global f0
    wb = xlrd.open_workbook(root_files_Global[t])
    sheet_names = wb.sheet_names()
    sheet = wb.sheet_by_name(sheet_names[0])
    for r in range(sheet.nrows):
        for c in range(sheet.ncols):
            va = sheet.cell(r, c).value
            row.append(r+f0)
            column.append(c)
            value.append(va)
    f0 = f0 + sheet.nrows
 
def read_excel_xlsx():#定义读取xlsx文件格式的函数，使用openpyxl
    global f0
    wb = openpyxl.load_workbook(root_files_Global[t])
    ws = wb.worksheets[0]
    for r in range(ws.max_row):
        for c in range(ws.max_column):
            va = ws.cell(r+1, c+1).value
            row.append(r + f0)
            column.append(c)
            value.append(va)
    f0 = f0 + ws.max_row
 
for t in range(len(root_files_Global)):
    if 'xlsx' in root_files_Global[t]:
        read_excel_xlsx()
    elif 'xls' in root_files_Global[t]:
        read_excel_xls()
 
def write_excel_xlsx():#定义写入xlsx文件格式的函数，使用openpyxl
    wb = openpyxl.Workbook()
    ws = wb.worksheets[0]
    for i in range(len(row)):
        ws.cell(row[i]+1,column[i]+1).value = value[i]
    wb.save('总表.xlsx')
write_excel_xlsx()




import openpyxl

def remove_short_and_keyword_rows(file_path: str):
    # 打开 Excel 文件
    wb = openpyxl.load_workbook(file_path)
    ws = wb.active

    # 定义关键字列表
    keywords = ['书记', '包村', '序号','合计','电话','名单','村名','台账','联系电话',"责任人"]

    # 遍历所有行，找到长度小于 5 且行中包含指定关键字的行
    rows_to_remove = []
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
        should_remove = True
        for cell in row:
            if (str(cell.value)) !='None':
                if not any(keyword  in str(cell.value) for keyword in keywords):
                    should_remove = False
                    break
        if should_remove:
            rows_to_remove.append(row)

    # 删除找到的行
    for row in rows_to_remove:
     
        ws.delete_rows(row[0].row, amount=1)

    # 保存并关闭 Excel 文件
    wb.save(file_path)
    wb.close()




remove_short_and_keyword_rows('总表.xlsx')

def remove_duplicate_rows(file_path: str):
    # 打开 Excel 文件
    wb = openpyxl.load_workbook(file_path)
    ws = wb.active
    
    # 定义变量，用于存储所有唯一的行和重复的行
    unique_rows = []
    duplicate_rows = []

    # 遍历所有行
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row):

        # 将行中所有的单元格的值拼接起来成为一个字符串，并用 strip 函数去掉首尾的空格
        row_values = [str(cell.value or '').strip() for cell in row]
        row_str = ''.join(row_values)

        # 如果当前行的字符串已经存在于 unique_rows 中，则说明当前行是重复的
        # 否则将当前行添加到 unique_rows 中
        if row_str in unique_rows:
            duplicate_rows.append(row)
        else:
            unique_rows.append(row_str)

    # 删除重复的行
    for row in duplicate_rows:
        ws.delete_rows(row[0].row, amount=1)

    # 保存并关闭 Excel 文件
    wb.save(file_path)
    wb.close()
remove_duplicate_rows('总表.xlsx')
print("ok")
